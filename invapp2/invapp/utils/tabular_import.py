"""Utilities for parsing tabular uploads (CSV/TSV/XLSX)."""

from __future__ import annotations

import csv
import io
import os
from typing import Iterable

from werkzeug.datastructures import FileStorage


class TabularImportError(ValueError):
    """Raised when tabular uploads cannot be parsed."""


def _rows_to_csv_text(rows: Iterable[Iterable[object]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(["" if cell is None else cell for cell in row])
    return output.getvalue()


def parse_tabular_upload(file_storage: FileStorage) -> str:
    """Return CSV text for a CSV, TSV, or XLSX upload."""

    if not file_storage or not file_storage.filename:
        raise TabularImportError("No file uploaded.")

    _, ext = os.path.splitext(file_storage.filename)
    ext = ext.lower()

    if ext == ".csv":
        try:
            return file_storage.stream.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise TabularImportError("CSV import files must be UTF-8 encoded.") from exc

    if ext == ".tsv":
        try:
            text = file_storage.stream.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise TabularImportError("TSV import files must be UTF-8 encoded.") from exc
        reader = csv.reader(io.StringIO(text), delimiter="\t")
        return _rows_to_csv_text(reader)

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency is required at runtime
            raise TabularImportError("XLSX uploads require openpyxl to be installed.") from exc

        data = file_storage.stream.read()
        workbook = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        return _rows_to_csv_text(sheet.iter_rows(values_only=True))

    raise TabularImportError("Unsupported file type. Upload a CSV, TSV, or XLSX file.")


def preview_csv_text(csv_text: str, max_rows: int = 50) -> tuple[list[str], list[list[str]]]:
    """Return headers and a preview of rows from CSV text."""

    reader = csv.reader(io.StringIO(csv_text))
    headers = next(reader, [])
    preview_rows: list[list[str]] = []
    for idx, row in enumerate(reader):
        if idx >= max_rows:
            break
        preview_rows.append(["" if cell is None else str(cell) for cell in row])
    return headers, preview_rows
