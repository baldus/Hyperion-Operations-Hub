"""Services for physical inventory workflows."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Iterable

from flask import current_app
from sqlalchemy import inspect
from sqlalchemy import String, Text, func

from invapp.extensions import db
from invapp.models import (
    InventoryCountLine,
    InventorySnapshot,
    InventorySnapshotLine,
    Item,
    Location,
    Movement,
)



@dataclass(frozen=True)
class SnapshotLineInput:
    item_id: int
    primary_match_text: str
    secondary_match_text: str | None
    system_total_qty: Decimal
    uom: str | None
    notes: str | None
    source_description_text: str | None


@dataclass(frozen=True)
class ReconciliationRow:
    item_id: int
    part_number: str
    description: str | None
    uom: str | None
    system_total_qty: Decimal
    counted_total_qty: Decimal | None
    variance: Decimal | None
    status: str


@dataclass(frozen=True)
class ImportData:
    headers: list[str]
    rows: list[dict[str, str]]
    preview_rows: list[dict[str, str]]
    normalized_headers: list[str]


@dataclass(frozen=True)
class NormalizationOptions:
    trim: bool = True
    case_insensitive: bool = True
    remove_spaces: bool = False
    remove_dashes: bool = False


@dataclass(frozen=True)
class ItemFieldOption:
    name: str
    type_label: str
    is_string: bool


@dataclass(frozen=True)
class MatchOutcome:
    item_id: int | None
    status: str
    matched_on: str | None
    reason: str
    primary_value: str
    secondary_value: str | None


@dataclass(frozen=True)
class MatchPreview:
    total_rows: int
    eligible_rows: int
    matched_rows: int
    matched_primary: int
    matched_secondary: int
    unmatched_rows: int
    ambiguous_rows: int
    empty_rows: int
    match_rate: float
    unmatched_examples: list[dict[str, str]]
    collision_count: int
    collision_examples: list[dict[str, object]]


def _normalize_header(header: str) -> str:
    return header.strip().lower()


def normalize_import_header(header: str) -> str:
    return _normalize_header(header).replace(" ", "_")


def _normalize_description(value: str) -> str:
    return " ".join(value.strip().lower().split())


def _normalize_part_number(value: str) -> str:
    return value.strip().upper().replace(" ", "")


def _normalize_part_number_strict(value: str) -> str:
    return _normalize_part_number(value).replace("-", "").replace("_", "")


def normalize_match_value(value: str, options: NormalizationOptions) -> str:
    normalized = value
    if options.trim:
        normalized = normalized.strip()
    if options.case_insensitive:
        normalized = normalized.lower()
    if options.remove_spaces:
        normalized = normalized.replace(" ", "")
    if options.remove_dashes:
        normalized = normalized.replace("-", "").replace("_", "")
    return normalized


ALLOWED_IMPORT_KEYS = {
    "Item ID",
    "Item Name",
    "Item Description",
    "Item Class",
    "Inactive",
    "Description for Purchases",
    "Item Note",
    "Quantity On Hand",
    "Quantity Needed",
    "Is Taxable",
    "Part Number",
    "UOM",
    "Notes",
}
MAX_FIELD_LEN = 2000
MAX_ROW_BYTES = 50_000
MAX_EXTRAS_KEYS = 50
TRUNCATE_PREVIEW_LEN = 500


def _truncate_string(value: str) -> object:
    if len(value) <= MAX_FIELD_LEN:
        return value
    return {
        "_truncated": True,
        "length": len(value),
        "preview": value[:TRUNCATE_PREVIEW_LEN],
    }


def _sanitize_value(value: object) -> object:
    if isinstance(value, str):
        return _truncate_string(value)
    if isinstance(value, (int, float, bool)) or value is None:
        return value
    if isinstance(value, dict):
        return {str(key): _sanitize_value(val) for key, val in value.items()}
    if isinstance(value, (list, tuple)):
        return [_sanitize_value(val) for val in value]
    return _truncate_string(str(value))


def normalize_import_row(raw_row: object) -> dict[str, object]:
    invalid_json = False
    if isinstance(raw_row, dict):
        base = raw_row
    elif isinstance(raw_row, str):
        try:
            parsed = json.loads(raw_row)
        except json.JSONDecodeError:
            invalid_json = True
            base = {"raw": raw_row}
        else:
            base = parsed if isinstance(parsed, dict) else {"raw": parsed}
    else:
        base = {"raw": raw_row}

    seen: set[str] = set()
    duplicates: dict[str, object] = {}
    blank_header_count = 0

    normalized: dict[str, object] = {}
    for key, value in dict(base).items():
        key_text = "" if key is None else str(key).strip()
        if not key_text:
            blank_header_count += 1
            continue
        if key_text in seen:
            suffix = 1
            duplicate_key = f"{key_text}__dup{suffix}"
            while duplicate_key in duplicates:
                suffix += 1
                duplicate_key = f"{key_text}__dup{suffix}"
            duplicates[duplicate_key] = _sanitize_value(value)
            continue
        seen.add(key_text)
        normalized[key_text] = _sanitize_value(value)

    core = {key: val for key, val in normalized.items() if key in ALLOWED_IMPORT_KEYS}
    extras = {
        key: val
        for key, val in normalized.items()
        if key not in ALLOWED_IMPORT_KEYS
    }
    extras.update(duplicates)
    unknown_header_count = len(extras)

    row_data: dict[str, object] = {
        **core,
        "_extras": extras,
        "_meta": {
            "invalid_json": invalid_json,
            "blank_header_count": blank_header_count,
            "unknown_header_count": unknown_header_count,
            "duplicate_header_count": len(duplicates),
        },
    }

    row_size_bytes = len(json.dumps(row_data, default=str))
    if row_size_bytes > MAX_ROW_BYTES:
        trimmed_extras_keys = sorted(extras.keys())[:MAX_EXTRAS_KEYS]
        trimmed_extras = {key: extras[key] for key in trimmed_extras_keys}
        row_data["_extras"] = trimmed_extras
        row_data["_meta"].update(
            {
                "row_data_compacted": True,
                "original_size_bytes": row_size_bytes,
                "extras_truncated_count": max(0, len(extras) - len(trimmed_extras)),
            }
        )
        row_size_bytes = len(json.dumps(row_data, default=str))
        if row_size_bytes > MAX_ROW_BYTES:
            reduced_keys = sorted(trimmed_extras.keys())[:5]
            row_data["_extras"] = {key: trimmed_extras[key] for key in reduced_keys}
            row_data["_meta"]["extras_truncated_count"] = max(
                0, len(trimmed_extras) - len(reduced_keys)
            )
            row_size_bytes = len(json.dumps(row_data, default=str))
        if row_size_bytes > MAX_ROW_BYTES:
            row_data["_extras"] = {}
            row_data["_meta"]["extras_truncated_count"] = len(extras)

    serialized = json.dumps(row_data, default=str)
    return json.loads(serialized)


def normalize_row_data(row_data: object) -> dict[str, object]:
    return normalize_import_row(row_data)


def get_import_issue_schema_signature() -> dict[str, dict[str, object]]:
    inspector = inspect(db.engine)
    columns = inspector.get_columns("inventory_snapshot_import_issue")
    signature: dict[str, dict[str, object]] = {}
    for column in columns:
        col_type = column["type"]
        signature[column["name"]] = {
            "type": str(col_type),
            "nullable": column.get("nullable"),
            "length": getattr(col_type, "length", None),
        }
    return signature


def is_import_issue_schema_valid(signature: dict[str, dict[str, object]]) -> bool:
    def _type_contains(column: str, needles: tuple[str, ...]) -> bool:
        entry = signature.get(column)
        if not entry:
            return False
        type_name = str(entry.get("type", "")).lower()
        return any(needle in type_name for needle in needles)

    row_data_ok = _type_contains("row_data", ("jsonb", "json", "text"))
    primary_ok = _type_contains("primary_value", ("text",))
    secondary_ok = _type_contains("secondary_value", ("text",))
    return row_data_ok and primary_ok and secondary_ok


def log_import_issue_schema_drift() -> bool:
    signature = get_import_issue_schema_signature()
    valid = is_import_issue_schema_valid(signature)
    if not valid:
        current_app.logger.error(
            "Import issue schema drift detected: %s. Run: alembic -c alembic.ini upgrade head",
            signature,
        )
    return valid


def get_item_match_field_options() -> list[ItemFieldOption]:
    options: list[ItemFieldOption] = []
    for column in Item.__table__.columns:
        name = column.name
        if "sku" in name.lower():
            continue
        column_type = type(column.type).__name__
        is_string = isinstance(column.type, (String, Text))
        options.append(
            ItemFieldOption(
                name=name,
                type_label=column_type,
                is_string=is_string,
            )
        )
    return options


def _safe_decimal(value: str) -> Decimal | None:
    if value is None:
        return None
    raw = value.strip()
    if not raw:
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _combine_notes(description: str | None, notes: str | None) -> str | None:
    parts: list[str] = []
    if description:
        parts.append(f"Description: {description}")
    if notes:
        parts.append(f"Notes: {notes}" if description else notes)
    return " | ".join(parts) if parts else None


def parse_import_bytes(
    filename: str,
    payload: bytes,
) -> tuple[ImportData | None, list[str]]:
    ext = (filename or "").lower().rsplit(".", maxsplit=1)
    if len(ext) == 1:
        return None, ["Upload a CSV, TSV, or XLSX file."]
    extension = ext[-1]

    if extension in {"csv", "tsv"}:
        delimiter = "\t" if extension == "tsv" else ","
        try:
            text = payload.decode("utf-8-sig")
        except UnicodeDecodeError:
            return None, ["File must be UTF-8 encoded."]
        return _parse_delimited(text, delimiter)
    if extension == "xlsx":
        return _parse_xlsx(payload)
    return None, ["Unsupported file type. Use CSV, TSV, or XLSX."]


def _parse_delimited(
    text: str,
    delimiter: str,
) -> tuple[ImportData | None, list[str]]:
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return None, ["The uploaded file does not include any rows."]
    headers = [str(header).strip() if header is not None else "" for header in rows[0]]
    if not any(headers):
        return None, ["The uploaded file does not include headers."]
    headers = _ensure_unique_headers(headers)
    data_rows = [
        _row_to_dict(headers, row)
        for row in rows[1:]
        if any(cell for cell in row)
    ]
    preview_rows = data_rows[:50]
    normalized_headers = [normalize_import_header(h) for h in headers]
    return (
        ImportData(
            headers=headers,
            rows=data_rows,
            preview_rows=preview_rows,
            normalized_headers=normalized_headers,
        ),
        [],
    )


def _parse_xlsx(payload: bytes) -> tuple[ImportData | None, list[str]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    except Exception:
        return None, ["Unable to read the XLSX file. Please re-export and try again."]
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([cell if cell is not None else "" for cell in row])
    if not rows:
        return None, ["The uploaded file does not include any rows."]
    headers = [str(header).strip() if header is not None else "" for header in rows[0]]
    if not any(headers):
        return None, ["The uploaded file does not include headers."]
    headers = _ensure_unique_headers(headers)
    data_rows = [
        _row_to_dict(headers, row)
        for row in rows[1:]
        if any(cell for cell in row)
    ]
    preview_rows = data_rows[:50]
    normalized_headers = [normalize_import_header(h) for h in headers]
    return (
        ImportData(
            headers=headers,
            rows=data_rows,
            preview_rows=preview_rows,
            normalized_headers=normalized_headers,
        ),
        [],
    )


def _ensure_unique_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    unique_headers: list[str] = []
    for header in headers:
        header_key = header or "column"
        if header_key not in counts:
            counts[header_key] = 0
            unique_headers.append(header_key)
        else:
            counts[header_key] += 1
            unique_headers.append(f"{header_key}_{counts[header_key]}")
    return unique_headers


def _row_to_dict(headers: list[str], row: list[object]) -> dict[str, str]:
    row_values = list(row) + [""] * (len(headers) - len(row))
    return {
        header: "" if value is None else str(value).strip()
        for header, value in zip(headers, row_values)
    }


def get_item_field_candidates() -> tuple[list[str], list[str]]:
    part_keywords = ("part", "number", "num", "item", "code", "pn", "mpn")
    desc_keywords = ("desc", "description", "name", "title")
    columns = [col.name for col in Item.__table__.columns]

    def _resolve_override(key: str) -> list[str]:
        override = current_app.config.get(key)
        if not override:
            return []
        if isinstance(override, (list, tuple, set)):
            values = [str(value).strip() for value in override]
        else:
            values = [value.strip() for value in str(override).split(",")]
        return [value for value in values if value and value in columns and value != "sku"]

    override_part = _resolve_override("PHYS_INV_ITEM_ID_FIELDS")
    override_desc = _resolve_override("PHYS_INV_DESC_FIELDS")
    if override_part:
        part_fields = override_part
    else:
        part_fields = [
            name
            for name in columns
            if name != "sku" and any(keyword in name.lower() for keyword in part_keywords)
        ]
    if override_desc:
        desc_fields = override_desc
    else:
        desc_fields = [
            name
            for name in columns
            if any(keyword in name.lower() for keyword in desc_keywords)
        ]
    return part_fields, desc_fields


def build_item_lookup() -> tuple[dict[str, list[int]], dict[str, list[int]], dict[str, list[int]], dict[int, list[str]], list[str], list[str]]:
    part_fields, desc_fields = get_item_field_candidates()
    raw_map: dict[str, set[int]] = {}
    normalized_map: dict[str, set[int]] = {}
    strict_map: dict[str, set[int]] = {}
    descriptions: dict[int, list[str]] = {}

    if not part_fields:
        return raw_map, normalized_map, strict_map, descriptions, part_fields, desc_fields

    for item in Item.query.all():
        for field in part_fields:
            value = getattr(item, field, None)
            if not value:
                continue
            raw_value = str(value).strip()
            if not raw_value:
                continue
            raw_map.setdefault(raw_value, set()).add(item.id)
            normalized_map.setdefault(_normalize_part_number(raw_value), set()).add(item.id)
            strict_map.setdefault(_normalize_part_number_strict(raw_value), set()).add(item.id)
        for field in desc_fields:
            value = getattr(item, field, None)
            if not value:
                continue
            descriptions.setdefault(item.id, []).append(_normalize_description(str(value)))

    raw_lookup = {key: sorted(value) for key, value in raw_map.items()}
    normalized_lookup = {key: sorted(value) for key, value in normalized_map.items()}
    strict_lookup = {key: sorted(value) for key, value in strict_map.items()}
    return raw_lookup, normalized_lookup, strict_lookup, descriptions, part_fields, desc_fields


def suggest_column_mappings(
    import_data: ImportData,
    raw_lookup: dict[str, list[int]],
    normalized_lookup: dict[str, list[int]],
    strict_lookup: dict[str, list[int]],
    desc_fields: list[str],
) -> dict[str, str | None]:
    headers = import_data.headers
    suggestions: dict[str, str | None] = {
        "part_number": None,
        "quantity": None,
        "description": None,
        "uom": None,
        "notes": None,
    }

    quantity_scores: dict[str, float] = {}
    part_scores: dict[str, float] = {}
    desc_scores: dict[str, float] = {}
    uom_scores: dict[str, float] = {}
    notes_scores: dict[str, float] = {}

    for header in headers:
        values = [row.get(header, "") for row in import_data.rows]
        non_empty = [value for value in values if value.strip()]
        total = len(values) or 1
        numeric = sum(1 for value in values if _safe_decimal(value) is not None)
        numeric_ratio = numeric / total
        quantity_scores[header] = numeric_ratio

        matches = 0
        for value in non_empty:
            if value in raw_lookup:
                matches += 1
                continue
            normalized = _normalize_part_number(value)
            if normalized in normalized_lookup or _normalize_part_number_strict(value) in strict_lookup:
                matches += 1
        part_scores[header] = matches / (len(non_empty) or 1)

        text_density = sum(
            1 for value in non_empty if any(char.isalpha() for char in value)
        ) / (len(non_empty) or 1)
        desc_scores[header] = text_density

        normalized_header = normalize_import_header(header)
        if "uom" in normalized_header or "unit" in normalized_header:
            uom_scores[header] = 1.0
        if "note" in normalized_header or "comment" in normalized_header:
            notes_scores[header] = 1.0

    if part_scores:
        suggestions["part_number"] = max(part_scores, key=part_scores.get)
    if quantity_scores:
        suggestions["quantity"] = max(quantity_scores, key=quantity_scores.get)
    if desc_fields:
        normalized_desc_fields = {normalize_import_header(field) for field in desc_fields}
        for header in headers:
            if normalize_import_header(header) in normalized_desc_fields:
                suggestions["description"] = header
                break
    if not suggestions["description"] and desc_scores:
        suggestions["description"] = max(desc_scores, key=desc_scores.get)
    if uom_scores:
        suggestions["uom"] = max(uom_scores, key=uom_scores.get)
    if notes_scores:
        suggestions["notes"] = max(notes_scores, key=notes_scores.get)

    return suggestions


def build_item_field_lookup(
    items: Iterable[Item],
    field: str,
    options: NormalizationOptions,
) -> dict[str, list[int]]:
    lookup: dict[str, list[int]] = {}
    for item in items:
        value = getattr(item, field, None)
        if value is None:
            continue
        normalized = normalize_match_value(str(value), options)
        if not normalized:
            continue
        lookup.setdefault(normalized, [])
        if item.id not in lookup[normalized]:
            lookup[normalized].append(item.id)
    return lookup


def _find_collision_examples(
    lookup: dict[str, list[int]],
    field: str,
    limit: int = 5,
) -> tuple[int, list[dict[str, object]]]:
    collisions = [
        {"field": field, "value": key, "count": len(ids)}
        for key, ids in lookup.items()
        if len(ids) > 1
    ]
    return len(collisions), collisions[:limit]


def match_rows(
    rows: list[dict[str, str]],
    primary_upload_col: str,
    primary_item_field: str,
    secondary_upload_col: str | None,
    secondary_item_field: str | None,
    options: NormalizationOptions,
    items: Iterable[Item],
) -> tuple[list[MatchOutcome], int, list[dict[str, object]]]:
    results: list[MatchOutcome] = []
    primary_lookup = build_item_field_lookup(items, primary_item_field, options)
    primary_collision_count, primary_examples = _find_collision_examples(
        primary_lookup, primary_item_field
    )
    collision_examples = list(primary_examples)
    collision_count = primary_collision_count

    secondary_lookup: dict[str, list[int]] = {}
    if secondary_upload_col and secondary_item_field:
        secondary_lookup = build_item_field_lookup(items, secondary_item_field, options)
        secondary_collision_count, secondary_examples = _find_collision_examples(
            secondary_lookup, secondary_item_field
        )
        collision_count += secondary_collision_count
        collision_examples.extend(secondary_examples)

    for row in rows:
        primary_raw = (row.get(primary_upload_col) or "").strip()
        primary_normalized = normalize_match_value(primary_raw, options)
        secondary_raw = (row.get(secondary_upload_col) or "").strip() if secondary_upload_col else ""
        secondary_normalized = (
            normalize_match_value(secondary_raw, options)
            if secondary_upload_col and secondary_item_field
            else ""
        )

        if not primary_normalized and not secondary_normalized:
            results.append(
                MatchOutcome(
                    item_id=None,
                    status="empty",
                    matched_on=None,
                    reason="empty value",
                    primary_value=primary_raw,
                    secondary_value=secondary_raw or None,
                )
            )
            continue

        primary_candidates = primary_lookup.get(primary_normalized, []) if primary_normalized else []
        if len(primary_candidates) == 1:
            results.append(
                MatchOutcome(
                    item_id=primary_candidates[0],
                    status="matched",
                    matched_on="primary",
                    reason="matched primary",
                    primary_value=primary_raw,
                    secondary_value=secondary_raw or None,
                )
            )
            continue

        if len(primary_candidates) > 1:
            matched_id = None
            if secondary_normalized:
                secondary_candidates = secondary_lookup.get(secondary_normalized, [])
                if len(secondary_candidates) == 1 and secondary_candidates[0] in primary_candidates:
                    matched_id = secondary_candidates[0]
            if matched_id is not None:
                results.append(
                    MatchOutcome(
                        item_id=matched_id,
                        status="matched",
                        matched_on="secondary",
                        reason="matched secondary",
                        primary_value=primary_raw,
                        secondary_value=secondary_raw or None,
                    )
                )
            else:
                results.append(
                    MatchOutcome(
                        item_id=None,
                        status="ambiguous",
                        matched_on=None,
                        reason="ambiguous",
                        primary_value=primary_raw,
                        secondary_value=secondary_raw or None,
                    )
                )
            continue

        if secondary_normalized:
            secondary_candidates = secondary_lookup.get(secondary_normalized, [])
            if len(secondary_candidates) == 1:
                results.append(
                    MatchOutcome(
                        item_id=secondary_candidates[0],
                        status="matched",
                        matched_on="secondary",
                        reason="matched secondary",
                        primary_value=primary_raw,
                        secondary_value=secondary_raw or None,
                    )
                )
            elif len(secondary_candidates) > 1:
                results.append(
                    MatchOutcome(
                        item_id=None,
                        status="ambiguous",
                        matched_on=None,
                        reason="ambiguous",
                        primary_value=primary_raw,
                        secondary_value=secondary_raw or None,
                    )
                )
            else:
                results.append(
                    MatchOutcome(
                        item_id=None,
                        status="unmatched",
                        matched_on=None,
                        reason="no match",
                        primary_value=primary_raw,
                        secondary_value=secondary_raw or None,
                    )
                )
        else:
            results.append(
                MatchOutcome(
                    item_id=None,
                    status="unmatched",
                    matched_on=None,
                    reason="no match",
                    primary_value=primary_raw,
                    secondary_value=secondary_raw or None,
                )
            )

    return results, collision_count, collision_examples


def summarize_match_preview(
    rows: list[dict[str, str]],
    matches: list[MatchOutcome],
    secondary_enabled: bool,
    collision_count: int,
    collision_examples: list[dict[str, object]],
) -> MatchPreview:
    matched_rows = sum(1 for match in matches if match.status == "matched")
    matched_primary = sum(
        1 for match in matches if match.status == "matched" and match.matched_on == "primary"
    )
    matched_secondary = sum(
        1 for match in matches if match.status == "matched" and match.matched_on == "secondary"
    )
    unmatched_rows = sum(1 for match in matches if match.status == "unmatched")
    ambiguous_rows = sum(1 for match in matches if match.status == "ambiguous")
    empty_rows = sum(1 for match in matches if match.status == "empty")
    eligible_rows = sum(
        1
        for match in matches
        if match.primary_value.strip()
        or (secondary_enabled and (match.secondary_value or "").strip())
    )
    match_rate = (matched_rows / eligible_rows * 100.0) if eligible_rows else 0.0

    unmatched_examples: list[dict[str, str]] = []
    for match in matches:
        if match.status == "matched":
            continue
        if len(unmatched_examples) >= 10:
            break
        unmatched_examples.append(
            {
                "primary_value": match.primary_value,
                "secondary_value": match.secondary_value or "",
                "reason": match.reason,
            }
        )

    return MatchPreview(
        total_rows=len(rows),
        eligible_rows=eligible_rows,
        matched_rows=matched_rows,
        matched_primary=matched_primary,
        matched_secondary=matched_secondary,
        unmatched_rows=unmatched_rows,
        ambiguous_rows=ambiguous_rows,
        empty_rows=empty_rows,
        match_rate=match_rate,
        unmatched_examples=unmatched_examples,
        collision_count=collision_count,
        collision_examples=collision_examples,
    )


def suggest_matching_upload_column(
    import_data: ImportData,
    items: Iterable[Item],
    item_field: str,
    options: NormalizationOptions,
) -> str | None:
    lookup = build_item_field_lookup(items, item_field, options)
    best_header = None
    best_score = 0.0
    for header in import_data.headers:
        values = [row.get(header, "") for row in import_data.rows]
        non_empty = [value for value in values if value.strip()]
        if not non_empty:
            continue
        matches = sum(
            1 for value in non_empty if normalize_match_value(value, options) in lookup
        )
        score = matches / len(non_empty)
        if score > best_score:
            best_score = score
            best_header = header
    return best_header


def apply_duplicate_strategy(
    rows: list[dict[str, str]],
    primary_col: str,
    secondary_col: str | None,
    qty_col: str,
    strategy: str,
    options: NormalizationOptions,
) -> tuple[list[dict[str, str]], int]:
    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        primary_value = normalize_match_value(row.get(primary_col, "") or "", options)
        secondary_value = (
            normalize_match_value(row.get(secondary_col, "") or "", options)
            if secondary_col
            else ""
        )
        key = f"{primary_value}::{secondary_value}"
        grouped.setdefault(key, []).append(row)

    duplicate_groups = sum(1 for group in grouped.values() if len(group) > 1)
    if strategy not in {"sum", "first", "last"}:
        strategy = "sum"

    merged_rows: list[dict[str, str]] = []
    for group in grouped.values():
        if strategy == "first":
            merged_rows.append(group[0])
            continue
        if strategy == "last":
            merged_rows.append(group[-1])
            continue
        total = Decimal("0")
        for row in group:
            qty = _safe_decimal(row.get(qty_col, ""))
            if qty is not None:
                total += qty
        merged = dict(group[0])
        merged[qty_col] = str(total)
        merged_rows.append(merged)

    return merged_rows, duplicate_groups


def build_snapshot_lines(
    rows: list[dict[str, str]],
    matches: list[MatchOutcome],
    primary_col: str,
    secondary_col: str | None,
    desc_col: str | None,
    qty_col: str,
    uom_col: str | None,
    notes_col: str | None,
) -> tuple[list[SnapshotLineInput], list[str]]:
    errors: list[str] = []
    snapshot_lines: list[SnapshotLineInput] = []
    for idx, (row, match) in enumerate(zip(rows, matches), start=1):
        if match.item_id is None:
            continue
        raw_qty = row.get(qty_col, "")
        qty = _safe_decimal(raw_qty)
        if qty is None:
            errors.append(f"Row {idx}: quantity '{raw_qty}' must be numeric.")
            continue
        uom = (row.get(uom_col, "") if uom_col else "").strip() or None
        desc_value = (row.get(desc_col, "") if desc_col else "").strip() or None
        notes_value = (row.get(notes_col, "") if notes_col else "").strip() or None
        combined_notes = _combine_notes(desc_value, notes_value)
        secondary_value = (row.get(secondary_col, "") if secondary_col else "").strip() or None
        snapshot_lines.append(
            SnapshotLineInput(
                item_id=match.item_id,
                primary_match_text=row.get(primary_col, "").strip(),
                secondary_match_text=secondary_value,
                system_total_qty=qty,
                uom=uom,
                notes=combined_notes,
                source_description_text=desc_value,
            )
        )
    return snapshot_lines, errors


def get_item_display_values(
    item: Item,
    part_fields: list[str],
    desc_fields: list[str],
) -> tuple[str, str]:
    part_value = ""
    for field in part_fields:
        value = getattr(item, field, None)
        if value:
            part_value = str(value)
            break
    desc_value = ""
    for field in desc_fields:
        value = getattr(item, field, None)
        if value:
            desc_value = str(value)
            break
    return part_value, desc_value


def get_known_item_location_pairs() -> list[tuple[int, int]]:
    pairs: set[tuple[int, int]] = set()

    movement_pairs = (
        db.session.query(Movement.item_id, Movement.location_id)
        .filter(Movement.item_id.isnot(None), Movement.location_id.isnot(None))
        .distinct()
        .all()
    )
    pairs.update(movement_pairs)

    location_fields = [
        Item.default_location_id,
        Item.secondary_location_id,
        Item.point_of_use_location_id,
    ]
    for field in location_fields:
        item_pairs = (
            db.session.query(Item.id, field)
            .filter(field.isnot(None))
            .distinct()
            .all()
        )
        pairs.update({(item_id, location_id) for item_id, location_id in item_pairs})

    return sorted(pairs)


def ensure_count_lines_for_snapshot(snapshot: InventorySnapshot) -> int:
    with db.session.no_autoflush:
        item_ids = [line.item_id for line in snapshot.lines]
    if not item_ids:
        return 0

    item_id_set = set(item_ids)
    known_pairs = {
        (item_id, location_id)
        for item_id, location_id in get_known_item_location_pairs()
        if item_id in item_id_set
    }
    if not known_pairs:
        return 0

    existing_pairs = {
        (line.item_id, line.location_id)
        for line in InventoryCountLine.query.filter_by(snapshot_id=snapshot.id).all()
    }

    to_create = known_pairs - existing_pairs
    if not to_create:
        return 0

    new_lines = [
        InventoryCountLine(
            snapshot_id=snapshot.id,
            item_id=item_id,
            location_id=location_id,
        )
        for item_id, location_id in to_create
    ]
    db.session.add_all(new_lines)
    return len(new_lines)


def build_reconciliation_rows(snapshot_id: int) -> list[ReconciliationRow]:
    snapshot_lines = (
        InventorySnapshotLine.query.filter_by(snapshot_id=snapshot_id)
        .join(Item)
        .all()
    )
    count_lines = InventoryCountLine.query.filter_by(snapshot_id=snapshot_id).all()
    part_fields, desc_fields = get_item_field_candidates()

    counts_by_item: dict[int, dict[str, object]] = {}
    for line in count_lines:
        entry = counts_by_item.setdefault(
            line.item_id,
            {
                "total": Decimal("0"),
                "has_lines": False,
                "has_values": False,
                "fully_counted": True,
            },
        )
        entry["has_lines"] = True
        if line.counted_qty is None:
            entry["fully_counted"] = False
            continue
        entry["has_values"] = True
        entry["total"] = entry["total"] + line.counted_qty

    rows: list[ReconciliationRow] = []
    for line in snapshot_lines:
        item = line.item
        part_number, description = get_item_display_values(item, part_fields, desc_fields)
        count_entry = counts_by_item.get(line.item_id)
        status = "UNCOUNTED"
        counted_total = None
        variance = None
        fully_counted = False

        if count_entry is None or not count_entry["has_lines"]:
            status = "UNLOCATED"
        else:
            fully_counted = bool(count_entry["fully_counted"])
            if count_entry["has_values"]:
                counted_total = count_entry["total"]
                variance = counted_total - line.system_total_qty
                if variance == 0 and fully_counted:
                    status = "MATCH"
                elif variance > 0:
                    status = "OVER"
                elif variance < 0:
                    status = "SHORT"
                else:
                    status = "PARTIAL"
            else:
                status = "UNCOUNTED"

        rows.append(
            ReconciliationRow(
                item_id=item.id,
                part_number=part_number,
                description=description or None,
                uom=line.uom or item.unit,
                system_total_qty=line.system_total_qty,
                counted_total_qty=counted_total,
                variance=variance,
                status=status,
            )
        )

    rows.sort(key=lambda row: row.part_number)
    return rows


def build_count_sheet_rows(snapshot_id: int) -> Iterable[dict[str, object]]:
    lines = (
        InventoryCountLine.query.filter_by(snapshot_id=snapshot_id)
        .join(Item)
        .join(InventoryCountLine.location)
        .order_by(Location.code, Item.id)
        .all()
    )
    snapshot_lines = {
        line.item_id: line
        for line in InventorySnapshotLine.query.filter_by(snapshot_id=snapshot_id).all()
    }
    part_fields, desc_fields = get_item_field_candidates()

    for line in lines:
        snapshot_line = snapshot_lines.get(line.item_id)
        part_number, description = get_item_display_values(
            line.item, part_fields, desc_fields
        )
        yield {
            "location_code": line.location.code,
            "location_description": line.location.description,
            "part_number": part_number,
            "description": description,
            "uom": (snapshot_line.uom if snapshot_line else None) or line.item.unit,
            "system_total_qty": snapshot_line.system_total_qty if snapshot_line else None,
            "counted_qty": line.counted_qty,
            "notes": line.notes,
        }


def summarize_snapshot(snapshot_id: int) -> dict[str, int]:
    line_total = (
        db.session.query(func.count(InventorySnapshotLine.id))
        .filter(InventorySnapshotLine.snapshot_id == snapshot_id)
        .scalar()
    )
    count_total = (
        db.session.query(func.count(InventoryCountLine.id))
        .filter(InventoryCountLine.snapshot_id == snapshot_id)
        .scalar()
    )
    return {
        "snapshot_lines": int(line_total or 0),
        "count_lines": int(count_total or 0),
    }
