from __future__ import annotations

import hashlib
import json
import logging
import tempfile
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from flask import current_app
from io import BytesIO


from invapp.extensions import db
from invapp.models import OpenOrderLine, OpenOrderLineSnapshot, OpenOrderUpload


COLUMN_VARIANTS = {
    "so_no": ("SO No", "SO #", "Sales Order", "Sales Order No"),
    "so_state": ("SO State", "Sales Order State"),
    "so_date": ("SO Date", "Order Date"),
    "ship_by": ("Ship By", "Ship Date"),
    "customer_id": ("Customer ID", "Cust ID"),
    "customer_name": ("Customer Name", "Customer"),
    "item_id": ("Item ID", "Item", "Item No"),
    "line_description": ("Line Description", "Description", "Line Desc"),
    "uom": ("U/M ID", "UOM", "Unit"),
    "qty_ordered": ("Qty Ordered", "Quantity Ordered"),
    "qty_shipped": ("Qty Shipped", "Quantity Shipped"),
    "qty_remaining": ("Qty Remaining", "Quantity Remaining", "Qty Open"),
    "unit_price": ("Unit Price", "Price"),
    "part_number": ("Part Number", "Part No", "Part"),
}

EXTERNAL_FIELDS = tuple(COLUMN_VARIANTS.keys())
DATE_FIELDS = {"so_date", "ship_by"}
NUMERIC_FIELDS = {"qty_ordered", "qty_shipped", "qty_remaining", "unit_price"}

DEFAULT_INTERNAL_STATUS = "UNREVIEWED"


@dataclass(frozen=True)
class OpenOrderDiff:
    current_rows: list[dict]
    previous_open_lines: dict[str, OpenOrderLine]
    new_keys: set[str]
    completed_keys: set[str]
    still_open_keys: set[str]
    changed_rows: list[dict]


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _normalize_string(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_key_component(value: object) -> str:
    return _normalize_string(value).upper()


def _parse_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(stripped, fmt).date()
            except ValueError:
                continue
    return None


def _parse_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    try:
        return int(Decimal(str(value)))
    except (ValueError, ArithmeticError):
        return None


def _parse_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (ValueError, ArithmeticError):
        return None


def _value_equals(field: str, left: object, right: object) -> bool:
    if left is None and right is None:
        return True
    if field in DATE_FIELDS:
        return _parse_date(left) == _parse_date(right)
    if field in NUMERIC_FIELDS:
        left_val = _parse_decimal(left)
        right_val = _parse_decimal(right)
        return left_val == right_val
    return _normalize_string(left) == _normalize_string(right)


def _extract_headers(header_row: Iterable[object]) -> dict[str, int]:
    normalized = {}
    for idx, raw in enumerate(header_row):
        key = _normalize_header(raw)
        if key:
            normalized[key] = idx
    mapping: dict[str, int] = {}
    for canonical, variants in COLUMN_VARIANTS.items():
        for variant in variants:
            normalized_key = _normalize_header(variant)
            if normalized_key in normalized:
                mapping[canonical] = normalized[normalized_key]
                break
    missing = [field for field in EXTERNAL_FIELDS if field not in mapping]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    return mapping


def _row_value(row: tuple, index: int | None) -> object:
    if index is None:
        return None
    if index >= len(row):
        return None
    return row[index]


def parse_open_orders(file_stream) -> list[dict]:
    if isinstance(file_stream, (bytes, bytearray)):
        file_stream = BytesIO(file_stream)
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
        raise ValueError("openpyxl is required to import Excel files.") from exc
    try:
        workbook = load_workbook(file_stream, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive
        raise ValueError("Unable to read Excel workbook.") from exc

    if not workbook.sheetnames:
        raise ValueError("Excel workbook contains no sheets.")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel workbook is empty.")

    header_row = rows[0]
    mapping = _extract_headers(header_row)

    parsed_rows: list[dict] = []
    for row in rows[1:]:
        if not row or all(cell in (None, "") for cell in row):
            continue

        so_state = _normalize_string(_row_value(row, mapping["so_state"]))
        qty_remaining = _parse_int(_row_value(row, mapping["qty_remaining"]))
        is_open = (qty_remaining or 0) > 0 or so_state.upper() == "OPEN"
        if not is_open:
            continue

        record = {
            "so_no": _normalize_string(_row_value(row, mapping["so_no"])),
            "so_state": so_state,
            "so_date": _parse_date(_row_value(row, mapping["so_date"])),
            "ship_by": _parse_date(_row_value(row, mapping["ship_by"])),
            "customer_id": _normalize_string(_row_value(row, mapping["customer_id"])),
            "customer_name": _normalize_string(_row_value(row, mapping["customer_name"])),
            "item_id": _normalize_string(_row_value(row, mapping["item_id"])),
            "line_description": _normalize_string(
                _row_value(row, mapping["line_description"])
            ),
            "uom": _normalize_string(_row_value(row, mapping["uom"])),
            "qty_ordered": _parse_int(_row_value(row, mapping["qty_ordered"])),
            "qty_shipped": _parse_int(_row_value(row, mapping["qty_shipped"])),
            "qty_remaining": qty_remaining,
            "unit_price": _parse_decimal(_row_value(row, mapping["unit_price"])),
            "part_number": _normalize_string(_row_value(row, mapping["part_number"])),
        }

        natural_components = (
            _normalize_key_component(record["so_no"]),
            _normalize_key_component(record["item_id"]),
            _normalize_key_component(record["line_description"]),
            _normalize_key_component(record["customer_id"]),
        )
        key_source = "|".join(natural_components)
        record["natural_key"] = hashlib.sha1(key_source.encode("utf-8")).hexdigest()
        parsed_rows.append(record)

    if not parsed_rows:
        raise ValueError("No open order lines were found in the upload.")

    return parsed_rows


def compute_open_order_diff(
    current_rows: list[dict],
    previous_open_lines: Iterable[OpenOrderLine],
) -> OpenOrderDiff:
    previous_map = {line.natural_key: line for line in previous_open_lines}
    current_map = {row["natural_key"]: row for row in current_rows}

    current_keys = set(current_map)
    previous_keys = set(previous_map)

    new_keys = current_keys - previous_keys
    completed_keys = previous_keys - current_keys
    still_open_keys = current_keys & previous_keys

    changed_rows: list[dict] = []
    for key in still_open_keys:
        current = current_map[key]
        previous = previous_map[key]
        changes = []
        for field in EXTERNAL_FIELDS:
            current_value = current.get(field)
            previous_value = getattr(previous, field)
            if not _value_equals(field, previous_value, current_value):
                changes.append(
                    {
                        "field": field,
                        "before": previous_value,
                        "after": current_value,
                    }
                )
        if changes:
            changed_rows.append(
                {
                    "current": current,
                    "previous": previous,
                    "changes": changes,
                }
            )

    return OpenOrderDiff(
        current_rows=current_rows,
        previous_open_lines=previous_map,
        new_keys=new_keys,
        completed_keys=completed_keys,
        still_open_keys=still_open_keys,
        changed_rows=changed_rows,
    )


def _apply_external_fields(line: OpenOrderLine, record: dict) -> None:
    for field in EXTERNAL_FIELDS:
        setattr(line, field, record.get(field))


def _serialize_snapshot_value(value: object) -> object:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _serialize_snapshot(record: dict) -> dict:
    return {key: _serialize_snapshot_value(value) for key, value in record.items()}


def _stage_directory() -> Path:
    base = Path(tempfile.gettempdir()) / "hyperion_open_orders"
    base.mkdir(parents=True, exist_ok=True)
    return base


def stage_open_orders_import(
    file_bytes: bytes,
    filename: str,
    previous_upload_id: int | None,
    notes: str | None,
) -> str:
    token = uuid.uuid4().hex
    staging_dir = _stage_directory()
    file_path = staging_dir / f"{token}.xlsx"
    meta_path = staging_dir / f"{token}.json"

    file_path.write_bytes(file_bytes)
    meta_path.write_text(
        json.dumps(
            {
                "filename": filename,
                "notes": notes or "",
                "previous_upload_id": previous_upload_id,
            }
        ),
        encoding="utf-8",
    )
    return token


def load_staged_open_orders(token: str) -> tuple[bytes, dict]:
    staging_dir = _stage_directory()
    file_path = staging_dir / f"{token}.xlsx"
    meta_path = staging_dir / f"{token}.json"

    if not file_path.exists() or not meta_path.exists():
        raise FileNotFoundError("Staged import could not be found.")

    metadata = json.loads(meta_path.read_text(encoding="utf-8"))
    return file_path.read_bytes(), metadata


def clear_staged_open_orders(token: str) -> None:
    staging_dir = _stage_directory()
    for suffix in (".xlsx", ".json"):
        path = staging_dir / f"{token}{suffix}"
        try:
            path.unlink()
        except FileNotFoundError:
            continue


def commit_open_orders_import(
    file_bytes: bytes,
    filename: str,
    user_id: int | None,
    previous_upload_id: int | None,
    notes: str | None,
) -> OpenOrderUpload:
    logger = current_app.logger if current_app else logging.getLogger(__name__)
    current_rows = parse_open_orders(file_bytes)

    previous_upload = None
    if previous_upload_id is not None:
        previous_upload = OpenOrderUpload.query.get(previous_upload_id)

    previous_open_lines: list[OpenOrderLine] = []
    if previous_upload:
        previous_open_lines = (
            OpenOrderLine.query.filter(
                OpenOrderLine.system_state != "COMPLETED",
                OpenOrderLine.last_seen_upload_id == previous_upload.id,
            )
            .all()
        )

    diff = compute_open_order_diff(current_rows, previous_open_lines)
    file_hash = hashlib.sha1(file_bytes).hexdigest()
    now = datetime.utcnow()

    upload = OpenOrderUpload(
        uploaded_by_user_id=user_id,
        source_filename=filename,
        file_hash=file_hash,
        notes=notes,
        uploaded_at=now,
    )
    db.session.add(upload)
    db.session.flush()

    current_map = {row["natural_key"]: row for row in current_rows}

    for key in diff.completed_keys:
        line = diff.previous_open_lines[key]
        line.system_state = "COMPLETED"
        line.completed_at = now
        line.completed_upload_id = upload.id

    for key in diff.still_open_keys:
        line = diff.previous_open_lines[key]
        record = current_map[key]
        _apply_external_fields(line, record)
        line.last_seen_upload_id = upload.id
        line.last_seen_at = now
        if line.system_state == "NEW":
            line.system_state = "OPEN"

    for key in diff.new_keys:
        record = current_map[key]
        existing = OpenOrderLine.query.filter_by(natural_key=key).first()
        if existing:
            if existing.system_state == "COMPLETED":
                existing.system_state = "REOPENED"
                existing.completed_at = None
                existing.completed_upload_id = None
            _apply_external_fields(existing, record)
            existing.last_seen_upload_id = upload.id
            existing.last_seen_at = now
            line = existing
        else:
            line = OpenOrderLine(
                natural_key=key,
                system_state="NEW",
                internal_status=DEFAULT_INTERNAL_STATUS,
                first_seen_upload_id=upload.id,
                last_seen_upload_id=upload.id,
                first_seen_at=now,
                last_seen_at=now,
            )
            _apply_external_fields(line, record)
            db.session.add(line)

        db.session.flush()

    for record in current_rows:
        line = OpenOrderLine.query.filter_by(natural_key=record["natural_key"]).first()
        if not line:
            continue
        snapshot = OpenOrderLineSnapshot(
            upload_id=upload.id,
            line_id=line.id,
            snapshot_json=_serialize_snapshot(record),
        )
        db.session.add(snapshot)

    db.session.commit()
    logger.info(
        "Committed open orders import %s with %s rows", upload.id, len(current_rows)
    )
    return upload
